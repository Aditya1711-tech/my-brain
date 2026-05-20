import io

from app.services.documents.types import RawExtraction


def parse_xlsx(file_bytes: bytes, filename: str) -> RawExtraction:
    """Extract text and structured sheet data from an XLSX file."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    text_parts: list[str] = []
    tables: list[list[list[str]]] = []
    structured: dict[str, list[list[str]]] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(cell) if cell is not None else "" for cell in row]
            if any(c for c in cells):
                rows.append(cells)

        if rows:
            # First row as header, rest as data
            header = " | ".join(rows[0])
            text_parts.append(f"Sheet: {sheet_name}\n{header}")
            tables.append(rows)
            structured[sheet_name] = rows

    wb.close()

    return RawExtraction(
        text="\n\n".join(text_parts),
        tables=tables,
        page_count=len(wb.sheetnames),
        structured=structured,
    )
